from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pdfplumber
import pypdfium2 as pdfium
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
YELLOW_FONT = Font(color="9C6500")
NO_FILL = PatternFill(fill_type=None)


@dataclass(frozen=True)
class StatementEntry:
    source_file: str
    bank_name: str
    sequence: str
    transaction_date: date
    direction: str
    amount: Decimal
    counterparty: str
    summary: str
    serial: str


@dataclass
class LedgerEntry:
    row: int
    bank_name: str
    transaction_date: date
    direction: str
    amount: Decimal
    summary: str


def normalize_text(value: object) -> str:
    text = re.sub(r"\s+", "", str(value or "")).strip()
    if not text:
        return text
    original_cjk = sum("\u4e00" <= char <= "\u9fff" for char in text)
    candidates = [text]
    for source_encoding in ("latin1", "cp1252"):
        for target_encoding in ("utf-8", "gb18030"):
            try:
                candidates.append(text.encode(source_encoding).decode(target_encoding))
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
    return max(
        candidates,
        key=lambda candidate: (
            sum("\u4e00" <= char <= "\u9fff" for char in candidate),
            -abs(len(candidate) - len(text)),
        ),
    ) if original_cjk == 0 else text


BANK_ALIASES = (
    ("工商银行", ("中国工商银行", "工商银行", "INDUSTRIALANDCOMMERCIALBANKOFCHINA")),
    ("建设银行", ("中国建设银行", "建设银行", "CHINACONSTRUCTIONBANK")),
    ("农业银行", ("中国农业银行", "农业银行", "农行", "AGRICULTURALBANKOFCHINA")),
    ("中国银行", ("中国银行", "中行", "BANKOFCHINA")),
    ("交通银行", ("交通银行", "BANKOFCOMMUNICATIONS")),
    ("邮储银行", ("中国邮政储蓄银行", "邮政储蓄银行", "邮储银行", "POSTALSAVINGSBANKOFCHINA")),
    ("浦发银行", ("上海浦东发展银行", "浦东发展银行", "浦发银行", "浦发", "SHANGHAIPUDONGDEVELOPMENTBANK")),
    ("招商银行", ("招商银行", "招行", "CHINAMERCHANTSBANK")),
    ("中信银行", ("中信银行", "CHINACITICBANK")),
    ("兴业银行", ("兴业银行", "INDUSTRIALBANK")),
    ("民生银行", ("中国民生银行", "民生银行", "CHINAMINSHENGBANK")),
    ("光大银行", ("中国光大银行", "光大银行", "CHINAEVERBRIGHTBANK")),
    ("华夏银行", ("华夏银行", "HUAXIABANK")),
    ("平安银行", ("平安银行", "PINGANBANK")),
    ("广发银行", ("广发银行", "广东发展银行", "CHINAGUANGFABANK")),
    ("浙商银行", ("浙商银行", "CHINAZHESHANGBANK")),
    ("北京银行", ("北京银行", "BANKOFBEIJING")),
    ("宁波银行", ("宁波银行", "BANKOFNINGBO")),
    ("上海银行", ("上海银行", "BANKOFSHANGHAI")),
    ("江苏银行", ("江苏银行", "BANKOFJIANGSU")),
)


def detect_bank_name(value: object) -> str:
    text = normalize_text(value).upper()
    for bank_name, aliases in BANK_ALIASES:
        if any(normalize_text(alias).upper() in text for alias in aliases):
            return bank_name
    return ""


def parse_money(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    cleaned = str(value).replace(",", "").replace("￥", "").strip()
    if not cleaned:
        return None
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"无法识别金额：{value}") from exc


def parse_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"无法识别日期：{value}")


def parse_embedded_date(value: object) -> date | None:
    if value is None:
        return None
    text = normalize_text(value)
    match = re.search(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}|20\d{6}", text)
    return parse_date(match.group()) if match else None


def find_column(header: list[object], names: tuple[str, ...]) -> int | None:
    normalized = [normalize_text(cell) for cell in header]
    for name in names:
        target = normalize_text(name)
        if target in normalized:
            return normalized.index(target)
    return None


def extract_boc_text_entries(pdf_path: Path) -> list[StatementEntry]:
    entries: list[StatementEntry] = []
    current: dict[str, object] | None = None
    bank_name = "中国银行"

    def flush() -> None:
        nonlocal current
        if current is None:
            return
        debit = parse_money(current["debit"])
        credit = parse_money(current["credit"])
        if bool(debit) != bool(credit):
            notes = normalize_text(current["notes"])
            entries.append(
                StatementEntry(
                    source_file=pdf_path.name,
                    bank_name=bank_name,
                    sequence=str(current["sequence"]),
                    transaction_date=current["transaction_date"],
                    direction="outflow" if debit else "inflow",
                    amount=debit or credit,
                    counterparty=notes.split("/")[0],
                    summary=normalize_text(current["summary"]),
                    serial=normalize_text(current["serial"]),
                )
            )
        current = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").splitlines():
                if not line.startswith("|"):
                    continue
                parts = line.split("|")
                if len(parts) < 12:
                    continue
                sequence = parts[1].strip()
                booking_date = parts[2].strip()
                if sequence.isdigit() and re.fullmatch(r"\d{6}", booking_date):
                    flush()
                    current = {
                        "sequence": sequence,
                        "transaction_date": datetime.strptime(booking_date, "%y%m%d").date(),
                        "summary": parts[6].strip(),
                        "debit": parts[7].strip(),
                        "credit": parts[8].strip(),
                        "serial": parts[10].strip(),
                        "notes": parts[11].strip(),
                    }
                elif current is not None and not sequence:
                    current["summary"] = f'{current["summary"]}{parts[6].strip()}'
                    current["notes"] = f'{current["notes"]}{parts[11].strip()}'
            flush()
    return entries


def extract_spdb_entries(pdf_path: Path, bank_name: str) -> list[StatementEntry]:
    entries: list[StatementEntry] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if len(table) < 3 or "交易日期" not in normalize_text(table[0][0] if table[0] else ""):
                    continue
                for row in table[2:]:
                    if not row or len(row) < 9:
                        continue
                    date_text = normalize_text(row[0])
                    if not re.fullmatch(r"\d{4}(?:/\d{2}/\d{2}|\d{4})", date_text):
                        continue
                    transaction_date = parse_date(date_text)
                    debit = parse_money(row[2])
                    credit = parse_money(row[3])
                    if transaction_date is None or bool(debit) == bool(credit):
                        continue
                    entries.append(
                        StatementEntry(
                            source_file=pdf_path.name,
                            bank_name=bank_name,
                            sequence=normalize_text(row[1]),
                            transaction_date=transaction_date,
                            direction="outflow" if debit else "inflow",
                            amount=debit or credit,
                            counterparty=normalize_text(row[6]),
                            summary=normalize_text(row[8] or row[7]),
                            serial=normalize_text(row[1]),
                        )
                    )
    return entries


def extract_abc_entries(pdf_path: Path, bank_name: str) -> list[StatementEntry]:
    entries: list[StatementEntry] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                header_index = next(
                    (
                        index
                        for index, row in enumerate(table[:5])
                        if row and any("交易时间" in normalize_text(cell) for cell in row)
                    ),
                    None,
                )
                if header_index is None:
                    continue
                for row_number, row in enumerate(table[header_index + 1 :], 1):
                    if not row or len(row) < 4:
                        continue
                    date_text = str(row[0] or "").split()[0]
                    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
                        continue
                    transaction_date = parse_date(date_text)
                    credit = parse_money(row[1])
                    debit = parse_money(row[2])
                    if transaction_date is None or bool(debit) == bool(credit):
                        continue
                    entries.append(
                        StatementEntry(
                            source_file=pdf_path.name,
                            bank_name=bank_name,
                            sequence=str(row_number),
                            transaction_date=transaction_date,
                            direction="outflow" if debit else "inflow",
                            amount=debit or credit,
                            counterparty=normalize_text(row[5]) if len(row) > 5 else "",
                            summary=normalize_text(row[7]) if len(row) > 7 else "",
                            serial=f"{pdf_path.name}:{row_number}",
                        )
                    )
    return entries


def extract_signed_text_entries(pdf_path: Path, bank_name: str) -> list[StatementEntry]:
    entries: list[StatementEntry] = []
    transaction_pattern = re.compile(
        r"^(?P<date>\d{8})\s+(?P<body>.*?)\s+"
        r"(?P<amount>[+-]?[\d,]+\.\d{2})\s+"
        r"(?P<balance>-?[\d,]+\.\d{2})(?:\s+(?P<counterparty>.*))?$"
    )
    numbered_pattern = re.compile(
        r"^(?P<sequence>\d+)\s+(?P<date>\d{8})\s+"
        r"(?P<amount>[+-][\d,]+\.\d{2})\s+"
        r"(?P<balance>-?[\d,]+\.\d{2})(?:\s+(?P<counterparty>.*))?$"
    )
    with pdfplumber.open(pdf_path) as pdf:
        row_number = 0
        for page in pdf.pages:
            for line in (page.extract_text() or "").splitlines():
                match = numbered_pattern.match(line.strip()) or transaction_pattern.match(line.strip())
                if not match:
                    continue
                row_number += 1
                amount_text = match.group("amount")
                signed_amount = parse_money(amount_text.lstrip("+-"))
                direction = "outflow" if amount_text.startswith("-") else "inflow"
                body = match.groupdict().get("body") or ""
                sequence = match.groupdict().get("sequence") or str(row_number)
                entries.append(
                    StatementEntry(
                        source_file=pdf_path.name,
                        bank_name=bank_name,
                        sequence=sequence,
                        transaction_date=parse_date(match.group("date")),
                        direction=direction,
                        amount=signed_amount,
                        counterparty=normalize_text(match.groupdict().get("counterparty")),
                        summary=normalize_text(body),
                        serial=f"{pdf_path.name}:{sequence}",
                    )
                )
    return entries


def find_tesseract() -> Path | None:
    candidates = (
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    command = shutil.which("tesseract")
    return Path(command) if command else None


def parse_ocr_money(words: list[tuple[int, str]], minimum_x: float, maximum_x: float, width: int) -> Decimal | None:
    text = "".join(
        value
        for left, value in sorted(words)
        if minimum_x * width <= left < maximum_x * width
    )
    text = text.upper().replace("O", "0").replace("，", ",").replace("。", ".")
    text = re.sub(r"\s+", "", text)
    matches = re.findall(r"\d[\d,]*\.\d{2}", text)
    return parse_money(matches[-1]) if matches else None


def load_ocr_cache(pdf_path: Path, cache_dir: Path) -> list[StatementEntry] | None:
    fingerprint = f"ccb-v2|{pdf_path.name}|{pdf_path.stat().st_size}|{pdf_path.stat().st_mtime_ns}"
    cache_path = cache_dir / f"{hashlib.sha256(fingerprint.encode('utf-8')).hexdigest()}.json"
    if not cache_path.exists():
        return None
    data = json.loads(cache_path.read_text(encoding="utf-8"))
    return [
        StatementEntry(
            source_file=item["source_file"],
            bank_name=item["bank_name"],
            sequence=item["sequence"],
            transaction_date=parse_date(item["transaction_date"]),
            direction=item["direction"],
            amount=parse_money(item["amount"]),
            counterparty=item["counterparty"],
            summary=item["summary"],
            serial=item["serial"],
        )
        for item in data
    ]


def save_ocr_cache(pdf_path: Path, cache_dir: Path, entries: list[StatementEntry]) -> None:
    cache_dir.mkdir(parents=True, exist_ok=True)
    fingerprint = f"ccb-v2|{pdf_path.name}|{pdf_path.stat().st_size}|{pdf_path.stat().st_mtime_ns}"
    cache_path = cache_dir / f"{hashlib.sha256(fingerprint.encode('utf-8')).hexdigest()}.json"
    data = [
        {
            "source_file": entry.source_file,
            "bank_name": entry.bank_name,
            "sequence": entry.sequence,
            "transaction_date": entry.transaction_date.isoformat(),
            "direction": entry.direction,
            "amount": str(entry.amount),
            "counterparty": entry.counterparty,
            "summary": entry.summary,
            "serial": entry.serial,
        }
        for entry in entries
    ]
    cache_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def extract_ccb_ocr_entries(pdf_path: Path, bank_name: str) -> list[StatementEntry]:
    tesseract = find_tesseract()
    tessdata_dir = Path(__file__).resolve().parent / "tessdata"
    if tesseract is None or not (tessdata_dir / "chi_sim.traineddata").exists():
        return []

    cache_dir = Path(__file__).resolve().parent / ".ocr_cache"
    cached = load_ocr_cache(pdf_path, cache_dir)
    if cached is not None:
        return cached

    entries: list[StatementEntry] = []
    document = pdfium.PdfDocument(str(pdf_path))
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        for page_index in range(len(document)):
            print(
                f"OCR识别：{pdf_path.name} 第 {page_index + 1}/{len(document)} 页",
                flush=True,
            )
            image = document[page_index].render(scale=3).to_pil().rotate(90, expand=True)
            image_path = temp_path / "page.png"
            output_base = temp_path / "ocr"
            image.save(image_path)
            command = [
                str(tesseract),
                str(image_path),
                str(output_base),
                "-l",
                "chi_sim+eng",
                "--tessdata-dir",
                str(tessdata_dir),
                "--psm",
                "6",
                "tsv",
            ]
            result = subprocess.run(command, capture_output=True, text=True, timeout=180)
            tsv_path = output_base.with_suffix(".tsv")
            if result.returncode != 0 or not tsv_path.exists():
                continue

            groups: dict[tuple[str, str, str], list[tuple[int, str]]] = defaultdict(list)
            with tsv_path.open(encoding="utf-8", errors="replace", newline="") as handle:
                for row in csv.DictReader(handle, delimiter="\t"):
                    value = (row.get("text") or "").strip()
                    if not value:
                        continue
                    key = (row["block_num"], row["par_num"], row["line_num"])
                    groups[key].append((int(row["left"]), value))

            width = image.width
            page_row = 0
            for words in groups.values():
                ordered = sorted(words)
                date_word = next(
                    (
                        value
                        for left, value in ordered
                        if left < width * 0.15 and re.fullmatch(r"20\d{6}", value)
                    ),
                    None,
                )
                if date_word is None:
                    continue
                debit = parse_ocr_money(ordered, 0.55, 0.63, width)
                credit = parse_ocr_money(ordered, 0.63, 0.71, width)
                if bool(debit) == bool(credit):
                    continue
                page_row += 1
                counterparty = normalize_text(
                    "".join(
                        value
                        for left, value in ordered
                        if width * 0.37 <= left < width * 0.55
                    )
                )
                summary = normalize_text(
                    "".join(
                        value
                        for left, value in ordered
                        if width * 0.23 <= left < width * 0.37
                    )
                )
                entries.append(
                    StatementEntry(
                        source_file=pdf_path.name,
                        bank_name=bank_name,
                        sequence=f"{page_index + 1}-{page_row}",
                        transaction_date=parse_date(date_word),
                        direction="outflow" if debit else "inflow",
                        amount=debit or credit,
                        counterparty=counterparty,
                        summary=summary,
                        serial=f"{pdf_path.name}:{page_index + 1}:{page_row}",
                    )
                )

    save_ocr_cache(pdf_path, cache_dir, entries)
    return entries


def extract_statement_entries(pdf_paths: list[Path]) -> tuple[list[StatementEntry], list[str]]:
    entries: list[StatementEntry] = []
    unrecognized_files: list[str] = []
    for pdf_path in pdf_paths:
        entry_count_before = len(entries)
        with pdfplumber.open(pdf_path) as pdf:
            first_page_text = pdf.pages[0].extract_text() if pdf.pages else ""
            bank_name = detect_bank_name(pdf_path.name) or detect_bank_name(first_page_text)
            for page in pdf.pages:
                for table in page.extract_tables():
                    if not table:
                        continue
                    header = table[0]
                    sequence_col = find_column(header, ("序号",))
                    date_col = find_column(header, ("交易日期", "会计日期"))
                    debit_col = find_column(header, ("借方发生额", "借方金额"))
                    credit_col = find_column(header, ("贷方发生额", "贷方金额"))
                    counterparty_col = find_column(header, ("对方户名", "对方单位名称"))
                    summary_col = find_column(header, ("摘要",))
                    serial_col = find_column(header, ("流水号",))
                    required = (date_col, debit_col, credit_col)
                    if any(column is None for column in required):
                        continue

                    for row in table[1:]:
                        if not row or date_col >= len(row):
                            continue
                        transaction_date = parse_date(row[date_col])
                        if transaction_date is None:
                            continue
                        debit = parse_money(row[debit_col]) if debit_col < len(row) else None
                        credit = parse_money(row[credit_col]) if credit_col < len(row) else None
                        if bool(debit) == bool(credit):
                            continue
                        direction = "outflow" if debit else "inflow"
                        amount = debit or credit
                        entries.append(
                            StatementEntry(
                                source_file=pdf_path.name,
                                bank_name=bank_name,
                                sequence=normalize_text(row[sequence_col]) if sequence_col is not None else "",
                                transaction_date=transaction_date,
                                direction=direction,
                                amount=amount,
                                counterparty=(
                                    normalize_text(row[counterparty_col])
                                    if counterparty_col is not None and counterparty_col < len(row)
                                    else ""
                                ),
                                summary=(
                                    normalize_text(row[summary_col])
                                    if summary_col is not None and summary_col < len(row)
                                    else ""
                                ),
                                serial=(
                                    normalize_text(row[serial_col])
                                    if serial_col is not None and serial_col < len(row)
                                    else ""
                                ),
                            )
                        )
        if len(entries) == entry_count_before:
            entries.extend(extract_boc_text_entries(pdf_path))
        if len(entries) == entry_count_before and bank_name == "浦发银行":
            entries.extend(extract_spdb_entries(pdf_path, bank_name))
        if len(entries) == entry_count_before and bank_name == "农业银行":
            entries.extend(extract_abc_entries(pdf_path, bank_name))
        if len(entries) == entry_count_before and bank_name in {"招商银行", "平安银行"}:
            entries.extend(extract_signed_text_entries(pdf_path, bank_name))
        if len(entries) == entry_count_before and bank_name == "建设银行":
            entries.extend(extract_ccb_ocr_entries(pdf_path, bank_name))
        if len(entries) == entry_count_before:
            unrecognized_files.append(pdf_path.name)
    if not entries:
        raise ValueError("PDF 中没有提取到可识别的银行流水表格")

    unique_entries: list[StatementEntry] = []
    seen: set[tuple[object, ...]] = set()
    for entry in entries:
        identity = (
            entry.bank_name,
            entry.serial or f"{entry.source_file}:{entry.sequence}",
            entry.transaction_date,
            entry.direction,
            entry.amount,
            entry.counterparty,
        )
        if identity not in seen:
            seen.add(identity)
            unique_entries.append(entry)
    return unique_entries, unrecognized_files


def find_ledger_sheet(workbook, requested_sheet: str | None):
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ValueError(f"找不到工作表：{requested_sheet}")
        return workbook[requested_sheet]

    for worksheet in workbook.worksheets:
        for row in range(1, min(worksheet.max_row, 20) + 1):
            values = {
                normalize_text(worksheet.cell(row, col).value)
                for col in range(1, min(worksheet.max_column, 100) + 1)
            }
            has_date = bool({"日期", "记账日期", "制单日期"} & values)
            split_amount_format = has_date and (
                {"借方金额", "贷方金额"}.issubset(values)
                or {"借方(本币)", "贷方(本币)"}.issubset(values)
                or {"借方", "贷方"}.issubset(values)
            )
            split_local_currency_format = {
                "日期",
                "借方(本币)",
                "贷方(本币)",
            }.issubset(values)
            combined_amount_format = (
                has_date
                and "金额" in values
                and bool({"借贷描述", "借贷标识"} & values)
            )
            amount_only_format = has_date and bool(
                {"金额", "原币金额"} & values
            )
            if split_amount_format or split_local_currency_format or combined_amount_format or amount_only_format:
                return worksheet
    raise ValueError("Excel 中找不到可识别的日期和金额列")


def find_excel_headers(worksheet) -> tuple[int, dict[str, int]]:
    aliases = {
        "date": ("日期", "记账日期", "制单日期"),
        "auxiliary": ("辅助项", "辅助信息"),
        "summary": ("摘要", "项目文本", "名称"),
        "ledger_debit": ("借方金额", "借方(本币)", "借方"),
        "ledger_credit": ("贷方金额", "贷方(本币)", "贷方"),
        "amount": ("金额", "原币金额"),
        "direction": ("借贷描述", "借贷标识"),
        "bank_account": (
            "科目描述",
            "科目名称",
            "核算维度",
            "账户说明",
            "帐户说明",
            "银行名称",
            "开户行",
        ),
        "statement_date": ("对账单日期",),
        "statement_amount": ("对账单金额",),
        "counterparty": ("对方单位名称", "对方户名"),
    }
    for row in range(1, min(worksheet.max_row, 20) + 1):
        header = [worksheet.cell(row, col).value for col in range(1, min(worksheet.max_column, 100) + 1)]
        columns = {key: find_column(header, names) for key, names in aliases.items()}
        split_amount_format = all(
            columns[key] is not None for key in ("date", "ledger_debit", "ledger_credit")
        )
        combined_amount_format = all(
            columns[key] is not None for key in ("date", "amount", "direction")
        )
        amount_only_format = columns["date"] is not None and columns["amount"] is not None
        if split_amount_format or combined_amount_format or amount_only_format:
            populated_columns = [index for index, value in enumerate(header) if value not in (None, "")]
            next_column = max(populated_columns) + 1
            for key in ("statement_date", "statement_amount", "counterparty"):
                if columns[key] is None:
                    columns[key] = next_column
                    worksheet.cell(row, next_column + 1).value = aliases[key][0]
                    next_column += 1
            return row, {key: column + 1 for key, column in columns.items() if column is not None}
    raise ValueError("Excel 中找不到有效表头")


def load_ledger_entries(worksheet, header_row: int, columns: dict[str, int]) -> list[LedgerEntry]:
    entries: list[LedgerEntry] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        transaction_date = parse_date(worksheet.cell(row, columns["date"]).value)
        if "auxiliary" in columns:
            transaction_date = (
                parse_embedded_date(worksheet.cell(row, columns["auxiliary"]).value)
                or transaction_date
            )
        if transaction_date is None:
            continue
        if "ledger_debit" in columns and "ledger_credit" in columns:
            ledger_debit = parse_money(worksheet.cell(row, columns["ledger_debit"]).value)
            ledger_credit = parse_money(worksheet.cell(row, columns["ledger_credit"]).value)
            if bool(ledger_debit) == bool(ledger_credit):
                continue
            direction = "inflow" if ledger_debit else "outflow"
            amount = ledger_debit or ledger_credit
        elif "amount" in columns and "direction" in columns:
            amount = parse_money(worksheet.cell(row, columns["amount"]).value)
            direction_text = normalize_text(worksheet.cell(row, columns["direction"]).value)
            if not amount or direction_text not in {"借方", "贷方", "S", "H", "借", "贷"}:
                continue
            direction = "inflow" if direction_text in {"借方", "S", "借"} else "outflow"
        elif "amount" in columns:
            amount = parse_money(worksheet.cell(row, columns["amount"]).value)
            if not amount:
                continue
            amount = abs(amount)
            direction = ""
        else:
            continue
        entries.append(
            LedgerEntry(
                row=row,
                bank_name=(
                    detect_bank_name(worksheet.cell(row, columns["bank_account"]).value)
                    if "bank_account" in columns
                    else ""
                ),
                transaction_date=transaction_date,
                direction=direction,
                amount=amount,
                summary=normalize_text(worksheet.cell(row, columns.get("summary", 0)).value) if columns.get("summary") else "",
            )
        )
    return entries


def choose_candidate(ledger: LedgerEntry, candidates: list[int], statements: list[StatementEntry]) -> int | None:
    if len(candidates) == 1:
        return candidates[0]
    text_matches = [
        index
        for index in candidates
        if statements[index].counterparty
        and (
            statements[index].counterparty in ledger.summary
            or ledger.summary in statements[index].counterparty
        )
    ]
    return text_matches[0] if len(text_matches) == 1 else None


def reset_reconciliation_cells(worksheet, header_row: int, columns: dict[str, int]) -> None:
    for row in range(header_row + 1, worksheet.max_row + 1):
        for key in ("statement_date", "statement_amount", "counterparty"):
            cell = worksheet.cell(row, columns[key])
            cell.value = None
            cell.fill = NO_FILL
            cell.font = Font(color="000000")


def write_match(
    worksheet,
    ledger: LedgerEntry,
    statement: StatementEntry,
    columns: dict[str, int],
    date_gap: int,
) -> None:
    values = {
        "statement_date": statement.transaction_date,
        "statement_amount": float(statement.amount),
        "counterparty": statement.counterparty,
    }
    for key, value in values.items():
        cell = worksheet.cell(ledger.row, columns[key])
        cell.value = value
        cell.fill = GREEN_FILL if date_gap == 0 else YELLOW_FILL
        cell.font = GREEN_FONT if date_gap == 0 else YELLOW_FONT
    worksheet.cell(ledger.row, columns["statement_date"]).number_format = "yyyy/m/d"
    worksheet.cell(ledger.row, columns["statement_amount"]).number_format = "#,##0.00"


def build_difference_sheet(workbook, ledger_unmatched, statement_unmatched, statements, unrecognized_files):
    sheet_name = "核对差异"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    headers = ["差异类型", "Excel行号", "Excel日期", "方向", "金额", "PDF候选信息"]
    for column, value in enumerate(headers, 1):
        cell = worksheet.cell(1, column, value)
        cell.fill = YELLOW_FILL
        cell.font = YELLOW_FONT

    row = 2
    for source_file in unrecognized_files:
        worksheet.cell(row, 1, "PDF未识别")
        worksheet.cell(row, 6, source_file)
        row += 1

    for ledger, candidate_indexes, reason in ledger_unmatched:
        candidate_text = "；".join(
            f"{statements[index].source_file} 序号{statements[index].sequence} "
            f"{statements[index].transaction_date} {statements[index].counterparty}"
            for index in candidate_indexes
        )
        values = [reason, ledger.row, ledger.transaction_date, ledger.direction, float(ledger.amount), candidate_text]
        for column, value in enumerate(values, 1):
            worksheet.cell(row, column, value)
        worksheet.cell(row, 3).number_format = "yyyy/m/d"
        worksheet.cell(row, 5).number_format = "#,##0.00"
        row += 1

    for index in statement_unmatched:
        statement = statements[index]
        values = [
            "PDF未匹配",
            "",
            statement.transaction_date,
            statement.direction,
            float(statement.amount),
            f"{statement.source_file} 序号{statement.sequence} {statement.counterparty} {statement.summary}",
        ]
        for column, value in enumerate(values, 1):
            worksheet.cell(row, column, value)
        worksheet.cell(row, 3).number_format = "yyyy/m/d"
        worksheet.cell(row, 5).number_format = "#,##0.00"
        row += 1

    widths = {"A": 16, "B": 12, "C": 14, "D": 12, "E": 16, "F": 60}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"


def reconcile(
    pdf_paths: list[Path],
    excel_path: Path,
    output_path: Path,
    sheet_name: str | None,
    date_tolerance: int,
) -> tuple[int, int, int, int, int, int]:
    statements, unrecognized_files = extract_statement_entries(pdf_paths)
    workbook = load_workbook(excel_path)
    worksheet = find_ledger_sheet(workbook, sheet_name)
    header_row, columns = find_excel_headers(worksheet)
    first_statement_date = min(entry.transaction_date for entry in statements) - timedelta(days=date_tolerance)
    last_statement_date = max(entry.transaction_date for entry in statements) + timedelta(days=date_tolerance)
    all_ledger_entries = load_ledger_entries(worksheet, header_row, columns)
    bank_ledger_entries = [entry for entry in all_ledger_entries if entry.bank_name]
    ledger_entries = [
        entry
        for entry in (bank_ledger_entries or all_ledger_entries)
        if first_statement_date <= entry.transaction_date <= last_statement_date
    ]
    reset_reconciliation_cells(worksheet, header_row, columns)

    index_by_amount: dict[tuple[str, str, Decimal], list[int]] = defaultdict(list)
    index_by_amount_without_direction: dict[tuple[str, Decimal], list[int]] = defaultdict(list)
    for index, statement in enumerate(statements):
        if statement.bank_name:
            index_by_amount[(statement.bank_name, statement.direction, statement.amount)].append(index)
            index_by_amount_without_direction[(statement.bank_name, statement.amount)].append(index)

    used_statements: set[int] = set()
    unresolved = set(range(len(ledger_entries)))
    assignments: dict[int, int] = {}

    while True:
        proposals: dict[int, list[tuple[int, int]]] = defaultdict(list)
        for ledger_index in unresolved:
            ledger = ledger_entries[ledger_index]
            candidate_pool = (
                index_by_amount.get(
                    (ledger.bank_name, ledger.direction, ledger.amount), []
                )
                if ledger.direction
                else index_by_amount_without_direction.get(
                    (ledger.bank_name, ledger.amount), []
                )
            )
            candidates = [
                index
                for index in candidate_pool
                if index not in used_statements
                and abs((statements[index].transaction_date - ledger.transaction_date).days) <= date_tolerance
            ]
            if not candidates:
                continue
            text_candidates = [
                index
                for index in candidates
                if statements[index].counterparty
                and statements[index].counterparty in ledger.summary
            ]
            if text_candidates:
                candidates = text_candidates
            minimum_gap = min(abs((statements[index].transaction_date - ledger.transaction_date).days) for index in candidates)
            nearest = [
                index
                for index in candidates
                if abs((statements[index].transaction_date - ledger.transaction_date).days) == minimum_gap
            ]
            if len(nearest) == 1:
                proposals[nearest[0]].append((ledger_index, minimum_gap))

        progress = False
        for statement_index, proposed_ledgers in proposals.items():
            minimum_gap = min(gap for _, gap in proposed_ledgers)
            nearest_ledgers = [item for item in proposed_ledgers if item[1] == minimum_gap]
            if len(nearest_ledgers) != 1:
                continue
            ledger_index, _ = nearest_ledgers[0]
            assignments[ledger_index] = statement_index
            unresolved.remove(ledger_index)
            used_statements.add(statement_index)
            progress = True
        if not progress:
            break

    exact_matches = 0
    tolerance_matches = 0
    for ledger_index, statement_index in assignments.items():
        ledger = ledger_entries[ledger_index]
        statement = statements[statement_index]
        date_gap = abs((statement.transaction_date - ledger.transaction_date).days)
        write_match(worksheet, ledger, statement, columns, date_gap)
        if date_gap == 0:
            exact_matches += 1
        else:
            tolerance_matches += 1

    unmatched_ledger = []
    for ledger_index in sorted(unresolved):
        ledger = ledger_entries[ledger_index]
        candidate_pool = (
            index_by_amount.get(
                (ledger.bank_name, ledger.direction, ledger.amount), []
            )
            if ledger.direction
            else index_by_amount_without_direction.get(
                (ledger.bank_name, ledger.amount), []
            )
        )
        amount_candidates = [
            index
            for index in candidate_pool
            if index not in used_statements
        ]
        tolerance_candidates = [
            index
            for index in amount_candidates
            if abs((statements[index].transaction_date - ledger.transaction_date).days) <= date_tolerance
        ]
        if tolerance_candidates:
            reason = "多笔候选待核对"
            difference_candidates = tolerance_candidates
        elif amount_candidates:
            reason = "日期超出容差"
            difference_candidates = amount_candidates
        else:
            reason = "Excel未匹配"
            difference_candidates = []
        unmatched_ledger.append((ledger, difference_candidates, reason))

    unmatched_statements = [index for index in range(len(statements)) if index not in used_statements]
    if "核对差异" in workbook.sheetnames:
        del workbook["核对差异"]
    workbook.save(output_path)
    difference_count = len(unrecognized_files) + len(unmatched_ledger) + len(unmatched_statements)
    unknown_bank_count = sum(1 for statement in statements if not statement.bank_name)
    return (
        len(statements),
        exact_matches,
        tolerance_matches,
        difference_count,
        len(unrecognized_files),
        unknown_bank_count,
    )


def find_default_file(directory: Path, suffix: str) -> Path:
    candidates = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() == suffix
        and not path.name.startswith("~$")
        and "_核对结果" not in path.stem
    ]
    if len(candidates) != 1:
        names = "、".join(path.name for path in candidates) or "无"
        raise ValueError(f"目录中需要恰好一个 {suffix} 文件，当前为：{names}")
    return candidates[0]


def find_pdf_files(directory: Path) -> list[Path]:
    candidates = sorted(
        path for path in directory.iterdir() if path.is_file() and path.suffix.lower() == ".pdf"
    )
    if not candidates:
        raise ValueError("目录中没有 PDF 文件")
    return candidates


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据日期、金额和收支方向核对银行 PDF 与 Excel")
    parser.add_argument("--pdf", type=Path, nargs="+", help="一个或多个银行对账单 PDF 路径")
    parser.add_argument("--excel", type=Path, help="序时账 Excel 路径")
    parser.add_argument("--output", type=Path, help="结果 Excel 路径")
    parser.add_argument("--sheet", help="指定工作表名称")
    parser.add_argument("--date-tolerance", type=int, default=62, help="允许的日期相差天数，默认 62 天")
    parser.add_argument("--overwrite", action="store_true", help="覆盖原 Excel")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    directory = Path.cwd()
    try:
        pdf_paths = [path.resolve() for path in args.pdf] if args.pdf else find_pdf_files(directory)
        excel_path = (args.excel or find_default_file(directory, ".xlsx")).resolve()
        if args.overwrite:
            output_path = excel_path
        elif args.output:
            output_path = args.output.resolve()
        else:
            output_path = excel_path.with_name(f"{excel_path.stem}_核对结果.xlsx")

        if args.date_tolerance < 0:
            raise ValueError("日期容差不能小于 0")
        (
            statement_count,
            exact_count,
            tolerance_count,
            difference_count,
            unrecognized_count,
            unknown_bank_count,
        ) = reconcile(
            pdf_paths, excel_path, output_path, args.sheet, args.date_tolerance
        )
        print(f"PDF文件：{len(pdf_paths)} 个")
        print(f"未识别PDF：{unrecognized_count} 个")
        print(f"银行未识别流水：{unknown_bank_count} 笔")
        print(f"PDF流水：{statement_count} 笔")
        print(f"精确匹配：{exact_count} 笔")
        print(f"容差匹配：{tolerance_count} 笔（±{args.date_tolerance} 天）")
        print(f"未匹配记录：{difference_count} 条")
        print(f"结果文件：{output_path}")
        return 0
    except Exception as exc:
        print(f"核对失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
